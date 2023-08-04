import PyPDF2

import openpyxl

import tkinter as tk

from tkinter import filedialog, messagebox, simpledialog

import re

from openpyxl.styles import Font, PatternFill

from openpyxl.utils import get_column_letter

from openpyxl.styles import numbers

import xlsxwriter

import openpyxl.styles as styles

from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED2

from openpyxl.styles.numbers import FORMAT_PERCENTAGE

import os

import pandas as pd

import math

from subprocess import call

from PIL import Image, ImageTk

import os 

from ReviewComposerV2.settings import MEDIA_ROOT 

print(MEDIA_ROOT)
MAIN_PATH = os.getcwd()

# This function is responsible for extracting the company name and date so that it can be stored in a txt file to then later be interacted with the API


def find_profit_loss_pages(pdf_path):

    profit_loss_pages = []

    with open(pdf_path, 'rb') as file:

        reader = PyPDF2.PdfReader(file)

        for page_number, page in enumerate(reader.pages, 1):

            extracted_text = page.extract_text()

            lines = extracted_text.strip().split(

                "\n")[:5]  # Limit to the first few lines

            for line in lines:

                # Search for "PROFIT AND LOSS ACCOUNT" in the line (case-insensitive)

                if re.search(r"PROFIT AND LOSS ACCOUNT", line, re.IGNORECASE):

                    profit_loss_pages.append(page_number)

                    break  # Found, no need to check further lines on this page

    return profit_loss_pages


def leap_year_checker(year):

    yr = int(year)

    if ((yr % 4 == 0 and yr % 100 != 0) or (yr % 400 == 0)):

        return True

    else:

        return False


def calculate_start_date(end_date):

    DICTIONARY = {

        'JANUARY': 'FEBRUARY',

        'FEBRUARY': 'MARCH',

        'MARCH': 'APRIL',

        'APRIL': 'MAY',

        'MAY': 'JUNE',

        'JUNE': 'JULY',

        'JULY': 'AUGUST',

        'AUGUST': 'SEPTEMBER',

        'SEPTEMBER': 'OCTOBER',

        'OCTOBER': 'NOVEMBER',

        'NOVEMBER': 'DECEMBER',

        'DECEMBER': 'JANUARY'

    }

    start_date = ['', '', '']

    end_date_pieces = end_date.split(" ")

    month_30th_end = ['APRIL', 'JUNE', 'SEPTEMBER', 'NOVEMBER']

    if end_date_pieces[1] != 'FEBRUARY':

        if end_date_pieces[1] not in month_30th_end:

            if end_date_pieces[0] == '31':

                start_date[0] = '01'

                start_date[1] = DICTIONARY[end_date_pieces[1]]

                start_date[2] = str(int(end_date_pieces[2]) - 1)

            else:

                start_date[0] = str(int(end_date_pieces[0]) - 1)

                start_date[1] = DICTIONARY[end_date_pieces[1]]

                start_date[2] = str(int(end_date_pieces[2]) - 1)

        else:

            if end_date_pieces[0] == '30':

                start_date[0] = '01'

                start_date[1] = DICTIONARY[end_date_pieces[1]]

                start_date[2] = str(int(end_date_pieces[2]) - 1)

            else:

                start_date[0] = str(int(end_date_pieces[0]) - 1)

                start_date[1] = DICTIONARY[end_date_pieces[1]]

                start_date[2] = str(int(end_date_pieces[2]) - 1)

    else:

        if leap_year_checker(end_date_pieces[2]):

            if end_date_pieces[0] == '29':

                start_date[0] = '01'

                start_date[1] = DICTIONARY[end_date_pieces[1]]

                start_date[2] = str(int(end_date_pieces[2]) - 1)

            else:

                start_date[0] = str(int(end_date_pieces[0]) - 1)

                start_date[1] = DICTIONARY[end_date_pieces[1]]

                start_date[2] = str(int(end_date_pieces[2]) - 1)

        else:

            if end_date_pieces[0] == '28':

                start_date[0] = '01'

                start_date[1] = DICTIONARY[end_date_pieces[1]]

                start_date[2] = str(int(end_date_pieces[2]) - 1)

            else:

                start_date[0] = str(int(end_date_pieces[0]) - 1)

                start_date[1] = DICTIONARY[end_date_pieces[1]]

                start_date[2] = str(int(end_date_pieces[2]) - 1)

    return " ".join(start_date)


def extract_company_and_date(pdf_path):

    with open(pdf_path, 'rb') as file:

        reader = PyPDF2.PdfReader(file)

        first_page = reader.pages[0]

        extracted_text = first_page.extract_text()

        # Split the extracted text into rows

        rows = extracted_text.split("\n")

        # Find the row index where it says "REGISTERED NUMBER"

        registered_number_row = None

        for row_index, row in enumerate(rows):

            if "REGISTERED NUMBER" in row.upper():

                registered_number_row = row_index

                break

        # Extract the company name from the row below "REGISTERED NUMBER" if available

        if registered_number_row is not None and registered_number_row < len(rows) - 1:

            company_name_row = registered_number_row + 1

            company_name = rows[company_name_row].strip()

        else:

            # If "REGISTERED NUMBER" row not found or it's the last row, take the top row on that page as the company name

            company_name = rows[0].strip()

        # Extract date from the bottom row

        date_pattern = r"(\d{1,2} [A-Za-z]+ \d{4})"

        date_matches = re.findall(date_pattern, extracted_text, re.MULTILINE)

        start_date = None

        end_date = None

        if "TO" in extracted_text:

            # Extract the date on the left side of "TO" as the start date

            start_date_match = re.search(






                r"(\d{1,2} [A-Za-z]+ \d{4}) TO", extracted_text, re.MULTILINE)

            if start_date_match:

                start_date = start_date_match.group(1).strip()

            # Extract the date to the right side of "TO" as the end date

            end_date_match = re.search(






                r"TO (\d{1,2} [A-Za-z]+ \d{4})", extracted_text, re.MULTILINE)

            if end_date_match:

                end_date = end_date_match.group(1).strip()

        if not start_date and len(date_matches) >= 1:

            # If "TO" is not present, assume the date is the end date

            end_date = date_matches[-1].strip()

        if start_date is None:

            start_date = calculate_start_date(end_date)

    return company_name, start_date, end_date


# This function utilizes the PyPDF2 module to extract data from the uploaded PDFs. This where it really starts.


def extract_data(pdf_path, page_numbers):

    with open(pdf_path, 'rb') as file:

        reader = PyPDF2.PdfReader(file)

        num_pages = len(reader.pages)

        for page_number in page_numbers:

            if page_number < 1 or page_number > num_pages:

                messagebox.showerror("Invalid Page Number",






                                     "Please enter valid page numbers.")

                return

        workbook = openpyxl.Workbook()

        sheet = workbook.active

        xero_sheet = workbook.create_sheet("Xero")

        xero_sheet.append(["Codes", "Expenditure", "Description",
                          "Sub Total", "Total Amount", "Date"])

        sheet.column_dimensions['A'].width = 30

        sheet.column_dimensions['B'].width = 20

        sheet.column_dimensions['C'].width = 20

        sheet.column_dimensions['D'].width = 20

        sheet.column_dimensions['E'].width = 20

        sheet.column_dimensions['F'].width = 50

        sheet.column_dimensions['G'].width = 30

        # sheet.merge_cells(start_row=4, start_column=7, end_row=8, end_column=7)

        font = Font(size=9, bold=True, underline="single")

        key_font = Font(size=9, bold=False)

        key_font_title = Font(size=11, bold=True)

        normal_font = Font(size=9)

        fill = PatternFill(start_color="B7CDE8",






                           end_color="B7CDE8", fill_type="solid")

        fillNETgreen = PatternFill(






            start_color="B7E8BA", end_color="B7E8BA", fill_type="solid")

        fillNETred = PatternFill(start_color="E8BAB7",






                                 end_color="E8BAB7", fill_type="solid")

        # The code below simply applies the font/styling to the column titles for each of the sheets

        column_titles = ["A", "B", "C", "D", "E", "F", "G"]

        for col_title in column_titles:

            sheet[col_title + '1'].font = font

        for col_title in column_titles:

            xero_sheet[col_title + '1'].font = font

       # Column title names

        sheet['A1'] = "Expenditure"

        sheet['B1'] = "This Year"

        sheet['C1'] = "Last Year"

        sheet['D1'] = "Difference"

        sheet['E1'] = "% Change"

        sheet['F1'] = "Explanation"

        sheet['G1'] = "Additional"

        sheet['I4'].value = "Key"

        sheet['H4'].font = Font(bold=True)

        # Cell H5: Colored cell with the same shade of blue

        cell_h5 = sheet['I5']

        cell_h5.fill = fill

        # Key messages

        sheet['I4'].font = key_font_title

        sheet['J5'].font = key_font

        sheet['J6'].font = key_font

        sheet['J5'].value = "Identified value 0 in this year's or last year's account."

        sheet['J6'].value = "Inaccurate calculations. Will need reviewing."

        cell_h5 = sheet['I8']

        cell_h5.fill = fillNETgreen

        # Key messages

        sheet['J9'].font = key_font

        sheet['J9'].value = "Positive figures compared to last year. (GROSS/NET PROFIT)"

        cell_h5 = sheet['I11']

        cell_h5.fill = fillNETred

        # Key messages

        sheet['J12'].font = key_font

        sheet['J12'].value = "Negative figures compared to last year. (GROSS/NET PROFIT)"

        row_index = 2  # Starting row index for data

        directors_sum = 0

        txt_list = []

        for i, page_number in enumerate(page_numbers, start=1):

            page = reader.pages[page_number - 1]

            extracted_text = page.extract_text()

            start_index = extracted_text.find("PROFIT AND LOSS ACCOUNT")

            end_index = extracted_text.find("Page")

            profit_loss_text = extracted_text[start_index:end_index].strip()

            pattern = r"^(.*?)\s+(-|\d[\d,.]*)\s+(-|\d[\d,.]*)$"

            data = []

            for line in profit_loss_text.split("\n"):

                match = re.match(pattern, line)

                if match:

                    expenditure = match.group(1)

                    this_year = match.group(2).replace(",", "")

                    last_year = match.group(3).replace(",", "")

                    if this_year == "-":

                        this_year_float = 0.0

                    else:

                        this_year_float = float(this_year)

                    if last_year == "-":

                        last_year_float = 0.0

                    else:

                        last_year_float = float(last_year)

                    difference = this_year_float - last_year_float

                    if last_year_float != 0.0:

                        percentage_change = (






                            difference / last_year_float) * 100

                    else:

                        percentage_change = 0.0

                    if difference > 0:

                        change_type = "increase"

                    elif difference < 0:

                        change_type = "decrease"

                    else:

                        change_type = "remained unchanged"

                    if this_year_float.is_integer():

                        this_year = int(this_year_float)

                    else:

                        this_year = round(this_year_float, 2)

                    if last_year_float.is_integer():

                        last_year = int(last_year_float)

                    else:

                        last_year = round(last_year_float, 2)

                    if difference.is_integer():

                        difference = int(difference)

                    else:

                        difference = round(difference, 2)

                    explanation = "£{:,} v £{:,} - £{:,} {} ({:.0f}%)".format(






                        this_year, last_year, difference, change_type, percentage_change)

                    if expenditure.startswith("Directors"):

                        directors_sum += this_year_float

                    data.append((expenditure, this_year, last_year,






                                difference, percentage_change, explanation))

                    BUILTIN_FORMATS = {






                        0: 'General',






                        1: '0',






                        2: '0.00',






                        3: '£#,##0',






                        4: '#,##0.00',






                        5: '0%'}

                    FORMAT_NUMBER_COMMA_SEPARATED1 = BUILTIN_FORMATS[3]

                    # FORMAT_PERCENTAGE = BUILTIN_FORMATS[5]

            for (expenditure, this_year, last_year, difference, percentage_change, explanation) in data:

                sheet.cell(row=row_index, column=1, value=expenditure)

                sheet.cell(row=row_index, column=2,






                           value=this_year).number_format = FORMAT_NUMBER_COMMA_SEPARATED1

                sheet.cell(row=row_index, column=3,






                           value=last_year).number_format = FORMAT_NUMBER_COMMA_SEPARATED1

                sheet.cell(row=row_index, column=4,






                           value=difference).number_format = FORMAT_NUMBER_COMMA_SEPARATED1

                sheet.cell(row=row_index, column=5,






                           value="{:.0f}%".format(percentage_change))

                sheet.cell(row=row_index, column=6,






                           value=explanation).number_format = FORMAT_NUMBER_COMMA_SEPARATED1

                for column_index in range(1, 7):

                    cell = sheet.cell(row=row_index, column=column_index)

                    cell.font = normal_font

                    if column_index > 6:

                        break  # Stop setting fill color after column F

                    if this_year == 0 or last_year == 0:

                        cell.fill = fill

                    if expenditure == "GROSS PROFIT" and this_year > last_year:

                        cell.fill = fillNETgreen

                    elif expenditure == "GROSS PROFIT" and this_year < last_year:

                        cell.fill = fillNETred

                    if expenditure == "NET PROFIT" and change_type == "increase":

                        cell.fill = fillNETgreen

                    elif expenditure == "NET PROFIT" and change_type == "decrease":

                        cell.fill = fillNETred

                row_index += 1

        directors_figures = []

        directors_found = False

        for row_index in range(2, row_index):

            expenditure = sheet.cell(row=row_index, column=1).value

            if expenditure and expenditure.startswith("Directors"):

                this_year = sheet.cell(row=row_index, column=2).value

                if this_year:

                    if isinstance(this_year, str):

                        this_year = float(this_year.replace(",", ""))

                    directors_figures.append(this_year)

        directors_sum = sum(directors_figures)

        # As long as NET PROFT is within 100 rows, this will locate it and store it

        net_figures = [0]

        for row_index in range(2, 100):

            expenditure = sheet.cell(row=row_index, column=1).value

            if expenditure and expenditure.startswith("NET"):

                this_year = sheet.cell(row=row_index, column=2).value

                if this_year:

                    if isinstance(this_year, str):

                        this_year = float(this_year.replace(",", ""))

                    net_figures.append(this_year)

        Net_sum = sum(net_figures)

        NT = directors_sum + Net_sum

        prev_directors_figures = []

        for row_index in range(2, row_index):

            expenditure = sheet.cell(row=row_index, column=1).value

            if expenditure and expenditure.startswith("Directors"):

                directors_found = True

                last_year = sheet.cell(row=row_index, column=3).value

                if last_year:

                    if isinstance(last_year, str):

                        last_year = float(last_year.replace(",", ""))

                    prev_directors_figures.append(last_year)

        prev_directors_sum = sum(prev_directors_figures)

        # As long as NET PROFT is within 100 rows, this will locate it and store it

        prev_net_figures = [0]

        for row_index in range(2, 100):

            expenditure = sheet.cell(row=row_index, column=1).value

            if expenditure and expenditure.startswith("NET"):

                last_year = sheet.cell(row=row_index, column=3).value

                if last_year:

                    if isinstance(last_year, str):

                        last_year = float(last_year.replace(",", ""))

                    prev_net_figures.append(last_year)

        prev_Net_sum = sum(prev_net_figures)

        prev_NT = prev_directors_sum + prev_Net_sum

        diff_NT = NT - prev_NT

        if prev_NT != 0.0:

            comp = (






                diff_NT / prev_NT) * 100

        else:

            comp = 0.0

        # comp = prev_NT / diff_NT * 100

        if diff_NT > 0:

            NTchange_type = "Increased"

        elif diff_NT < 0:

            NTchange_type = "Decreased"

        else:

            NTchange_type = "Remained unchanged"

        if directors_found:

            sheet['G7'].value = "Profit was £{:,} before Director remuneration. Compared to last year: £{:,}. {} by ({:.0f}%).".format(

                NT, prev_NT, NTchange_type, comp)

        else:

            sheet['G7'].value = "Director remuneration was not detected in the account uploaded."

        # sheet['G8'].value = "Compared to last year: £{:.0f} before Director remuneration.".format(prev_NT)

        # sheet['G9'].value = "{} by ({:.0f}%) from last to this year.".format(NTchange_type, comp)

        # Set the alignment and text wrap for the cell

        cell = sheet['G7']

        cell.alignment = styles.Alignment(






            horizontal='left', vertical='center', wrap_text=True)

        # Adjust the column width to fit the content

        sheet.column_dimensions['G'].width = 40

        sheet['G7'].font = normal_font

        sheet['G8'].font = normal_font

        sheet['G9'].font = normal_font

        start_cell = sheet['G7']

        end_cell = sheet['G9']

        merge_range = f'{start_cell.coordinate}:{end_cell.coordinate}'

        sheet.merge_cells(merge_range)

    return workbook


def file_asker():

    pdf_path = filedialog.askopenfilename(






        filetypes=[("PDF Files", "*.pdf")], title="Select PDF File")

    if pdf_path:

        company_name, start_date, end_date = extract_company_and_date(pdf_path)

        print(f"Company Name: {company_name}")

        print(f"Date: {start_date}")

        print(f"Date: {end_date}")

        # Prompt the user to enter the page numbers

        page_numbers = find_profit_loss_pages(pdf_path)

        workbook = extract_data(pdf_path, page_numbers)

        if workbook:

            # Prompt the user to choose the save location for the Excel file

            excel_path = filedialog.asksaveasfilename(defaultextension=".xlsx",












                                                      filetypes=[

                                                          ("Excel Files", "*.xlsx")],

                                                      title="Save Excel File")

            if excel_path:

                workbook.save(excel_path)

                output = "output"

                messagebox.showinfo("Extraction Complete",






                                    "Data extracted and saved to Excel file.")

                # Create path for text file using the excel path

                txt_path = os.path.join(MAIN_PATH,"pythonCollection","output.csv")

                xl_file = pd.read_excel(excel_path)

                xl_file['diff_float_pers'] = xl_file['Difference'] / \
                    xl_file['Last Year']

                new_df = xl_file[xl_file['diff_float_pers'] >= 0.05]

                final_df = new_df.copy()[['Expenditure', 'This Year']]

                final_df.to_csv(txt_path, index=False)

                with open(txt_path, "a") as file:

                    # Writing the company name and date into the txt file for the API to then recognise

                    file.write(f"{company_name}, 1.1\n")

                    file.write(f"{start_date}, 1.1\n")

                    file.write(f"{end_date}, 1.1\n")

                from subprocess import call

                call([os.path.join(MAIN_PATH,"pythonCollection","FINAL.py")])

                # 1. ) Invoice DF

                df_inv = pd.read_csv(os.path.join(MAIN_PATH,"pythonCollection",'invoice_new.csv'))

                codes = pd.DataFrame(df_inv['AccountCode'].value_counts())

                codes['acc_codes'] = list(codes.index)

                codes.index = range(codes.shape[0])

                codes.columns = ['no_of_invoices', 'acc_code']

                FRAMES_INV = {}

                for i in codes['acc_code']:

                    FRAMES_INV[i] = df_inv[(df_inv['AccountCode'] == i) & (

                        df_inv['Status'] == 'PAID')].sort_values(by="TotalAmount", ascending=False).head(10)

                # 2.) Transaction DF

                df_trans = pd.read_csv(os.path.join(MAIN_PATH,"pythonCollection",'transaction_new.csv'))

                codes_trans = pd.DataFrame(

                    df_trans['AccountCode'].value_counts())

                codes_trans['acc_codes'] = list(codes_trans.index)

                codes_trans.index = range(codes_trans.shape[0])

                codes_trans.columns = ['no_of_transacs', 'acc_code']

                FRAMES_TRANS = {}

                for i in codes_trans['acc_code']:

                    if str(i).isnumeric():

                        FRAMES_TRANS[int(i)] = df_trans[(df_trans['AccountCode'] == i) & (

                            df_trans['Status'] == 'AUTHORISED')].sort_values(by="TotalAmount", ascending=False).head(10)

                    else:

                        FRAMES_TRANS[i] = df_trans[(df_trans['AccountCode'] == i) & (

                            df_trans['Status'] == 'AUTHORISED')].sort_values(by="TotalAmount", ascending=False).head(10)

                # 3.) Overpayments DF

                df_over = pd.read_csv(os.path.join(MAIN_PATH,"pythonCollection",'overpayment_new.csv'))

                codes_over = pd.DataFrame(

                    df_over['AccountCode'].value_counts())

                codes_over['acc_codes'] = list(codes_over.index)

                codes_over.index = range(codes_over.shape[0])

                codes_over.columns = ['no_of_overpayments', 'acc_code']

                FRAMES_OVER = {}

                for i in codes_over['acc_code']:

                    if str(i).isnumeric():

                        FRAMES_OVER[int(i)] = df_over[(df_over['AccountCode'] == i)].sort_values(

                            by="TotalAmount", ascending=False).head(10)

                    else:

                        FRAMES_OVER[i] = df_over[(df_over['AccountCode'] == i)].sort_values(

                            by="TotalAmount", ascending=False).head(10)

                # 4. ) Jounals DF

                df_journal = pd.read_csv(os.path.join(MAIN_PATH,"pythonCollection",'journal_new.csv'))

                codes_journal = pd.DataFrame(

                    df_journal['AccountCode'].value_counts())

                codes_journal['acc_codes'] = list(codes_journal.index)

                codes_journal.index = range(codes_journal.shape[0])

                codes_journal.columns = ['no_of_journals', 'acc_code']

                FRAMES_JOURNAL = {}

                for i in codes_journal['acc_code']:

                    if str(i).isnumeric():

                        FRAMES_JOURNAL[int(i)] = df_journal[(df_journal['AccountCode'] == i) & (

                            df_journal['Status'] == 'POSTED')].sort_values(by="LineAmount", ascending=False).head(10)

                    else:

                        FRAMES_JOURNAL[i] = df_journal[(df_journal['AccountCode'] == i) & (

                            df_journal['Status'] == 'POSTED')].sort_values(by="LineAmount", ascending=False).head(10)

                # 5. )Prepaymants

                df_prepay = pd.read_csv(os.path.join(MAIN_PATH,"pythonCollection",'prepayment_new.csv'))

                codes_prepay = pd.DataFrame(

                    df_prepay['AccountCode'].value_counts())

                codes_prepay['acc_codes'] = list(codes_prepay.index)

                codes_prepay.index = range(codes_prepay.shape[0])

                codes_prepay.columns = ['no_of_transacs', 'acc_code']

                FRAMES_PREPAY = {}

                for i in codes_prepay['acc_code']:

                    if str(i).isnumeric():

                        FRAMES_PREPAY[int(i)] = df_prepay[(df_prepay['AccountCode'] == i) & (

                            df_prepay['Status'] == 'PAID')].sort_values(by="Total", ascending=False).head(10)

                    else:

                        FRAMES_PREPAY[i] = df_prepay[(df_prepay['AccountCode'] == i) & (

                            df_prepay['Status'] == 'PAID')].sort_values(by="Total", ascending=False).head(10)

                # 6. )PurchaseOrder

                df_order = pd.read_csv(os.path.join(MAIN_PATH,"pythonCollection",'purchaseorder_new.csv'))

                codes_order = pd.DataFrame(

                    df_order['AccountCode'].value_counts())

                codes_order['acc_codes'] = list(codes_order.index)

                codes_order.index = range(codes_order.shape[0])

                codes_order.columns = ['no_of_purchaseOrders', 'acc_code']

                FRAMES_ORDER = {}

                for i in codes_order['acc_code']:

                    if str(i).isnumeric():

                        FRAMES_ORDER[int(i)] = df_order[(df_order['AccountCode'] == i) & (

                            df_order['Status'] != 'DELETED')].sort_values(by="Total", ascending=False).head(10)

                    else:

                        FRAMES_ORDER[i] = df_order[(df_order['AccountCode'] == i) & (

                            df_order['Status'] != 'DELETED')].sort_values(by="Total", ascending=False).head(10)

                # 7. )CreditNotes

                df_notes = pd.read_csv(os.path.join(MAIN_PATH,"pythonCollection",'creditnotes_new.csv'))

                codes_notes = pd.DataFrame(

                    df_notes['AccountCode'].value_counts())

                codes_notes['acc_codes'] = list(codes_notes.index)

                codes_notes.index = range(codes_notes.shape[0])

                codes_notes.columns = ['no_of_creditNotes', 'acc_code']

                FRAMES_NOTES = {}

                for i in codes_notes['acc_code']:

                    if str(i).isnumeric():

                        FRAMES_NOTES[int(i)] = df_notes[(df_notes['AccountCode'] == i) & (

                            df_notes['Status'] == 'PAID')].sort_values(by="Total", ascending=False).head(10)

                    else:

                        FRAMES_NOTES[i] = df_notes[(df_notes['AccountCode'] == i) & (

                            df_notes['Status'] == 'PAID')].sort_values(by="Total", ascending=False).head(10)


def closeWindow():

    root.destroy()


########################################
root = tk.Tk()

root.config(background='#212121')

root.geometry('1050x575')

root.maxsize(1050, 800)

root.title("Review Notes Composer")

#asdadasdadadasdasdasdadaasdadadasd  ssssssssssssssssssss
ico = Image.open('logo-only.png')

photo = ImageTk.PhotoImage(ico)

root.wm_iconphoto(False, photo)

########################################


# Empty Row 1

empR1 = tk.Label(text='', background='#212121', width=1, height=5)

empR1.grid(row=0, column=0, columnspan=8)


# Empty Column 1

empC1 = tk.Label(text='', background='#212121', width=9)

empC1.grid(row=0, column=0, rowspan=8)


# Empty Column 2

empC2 = tk.Label(text='', background='#212121', width=1)

empC2.grid(row=0, column=1, rowspan=8)


# Warning Label

infoLabel = tk.Label(text=''' BETA VERSION  

   

                     \n1) Please use this software as an auxiliary tool to aid in existing processes. It is not quite a replacement yet.

2) This software can make errors. Ensure all the figures are correctly depicted in your work.

3) Please feel free to report any features you would like to see integrated or improved.

\n

Status: Currently looking into ways to utilise the XERO API in order to provide more intricate and robust notes.

\n

Important note: The page numbers requested are not the numbers at the bottom of the page. They are the PDF page numbers.

\n

Process:

1) Choose the PDF.

2) Input the PDF page numbers.

3) Choose where you want to save the newly formed excel sheet.

4) Save. Open. Review.

'''

                     )

# infoLabel.config(background='#f5f5dc',padx=5)

infoLabel.config(bg='#f5f5dc', font=(

    'MS Reference Sans Serif', 10), borderwidth=3, padx=5)

infoLabel.grid(row=1, column=2, columnspan=4)


# Empty row 2 (between warning label and compose review button)

empR2 = tk.Label(text='', background='#212121', width=1)

empR2.grid(row=3, column=0, columnspan=8)


# Compose notes button

start_btn = tk.Button(root, text='Compose review notes', command=file_asker)

# start_btn.config(padx=30,pady=10)

start_btn.config(padx=30, pady=10, background='#cccccc', fg='#000000', font=(

    'MS Reference Sans Serif', 10), borderwidth=3)


start_btn.grid(row=4, column=3, columnspan=2)


# empty row 3 (between compose notes button and cancel opeartion button)

empR3 = tk.Label(text='', background='#212121', width=1)

empR3.grid(row=5, column=0, columnspan=8)


# cancel button

cancelBtn = tk.Button(text='Cancel', command=closeWindow)

# cancelBtn.config(padx=15,pady=7)

cancelBtn.config(padx=10, pady=5, background='#cccccc', fg='#000000', font=(

    'MS Reference Sans Serif', 10), borderwidth=3)

cancelBtn.grid(row=6, column=6)


# Empty row 4 (at the end of the grid)

empR4 = tk.Label(text='', background='#212121', width=1)

empR4.grid(row=7, column=0, columnspan=8)


# Empty Column 3

empC3 = tk.Label(text='', background='#212121', width=1)

empC3.grid(row=0, column=7, rowspan=8)


root.mainloop()
